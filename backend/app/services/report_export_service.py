"""
Report Export Service
Handles exporting reports to PDF, Excel, and CSV formats
REQ-REPORT-012: Export reports to PDF, Excel, CSV
"""

from typing import Optional, Dict, Any
from datetime import datetime, timedelta
import uuid
import io
import csv
import json

from app.schemas.custom_reports import (
    GeneratedReportResponse,
    ReportExportResponse,
    ReportSectionData,
    ReportDataPoint
)


class ReportExportService:
    """Service for exporting reports to various formats"""

    def __init__(self):
        # In production, configure export storage location
        self.exports_cache: Dict[str, Any] = {}

    async def export_report(
        self,
        generated_report: GeneratedReportResponse,
        format: str,
        include_charts: bool = True,
        include_raw_data: bool = False
    ) -> ReportExportResponse:
        """
        Export report to specified format

        Args:
            generated_report: Generated report data
            format: Export format (pdf, excel, csv)
            include_charts: Whether to include visualizations
            include_raw_data: Whether to include raw data tables

        Returns:
            Export response with download URL
        """
        export_id = str(uuid.uuid4())

        if format == "pdf":
            file_data = await self._export_to_pdf(
                generated_report, include_charts, include_raw_data
            )
            filename = f"{generated_report.name.replace(' ', '_')}_report.pdf"
            content_type = "application/pdf"

        elif format == "excel":
            file_data = await self._export_to_excel(
                generated_report, include_charts, include_raw_data
            )
            filename = f"{generated_report.name.replace(' ', '_')}_report.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        elif format == "csv":
            file_data = await self._export_to_csv(generated_report)
            filename = f"{generated_report.name.replace(' ', '_')}_report.csv"
            content_type = "text/csv"

        else:
            raise ValueError(f"Unsupported format: {format}")

        # Cache the export
        self.exports_cache[export_id] = {
            "data": file_data,
            "content_type": content_type,
            "filename": filename,
            "created_at": datetime.now()
        }

        # Generate download URL
        download_url = f"/api/v1/reports/custom/export/{export_id}/download"

        return ReportExportResponse(
            export_id=export_id,
            report_id=generated_report.report_id,
            format=format,
            filename=filename,
            file_size=len(file_data),
            download_url=download_url,
            expires_at=datetime.now() + timedelta(hours=24),
            generated_at=datetime.now()
        )

    async def _export_to_pdf(
        self,
        report: GeneratedReportResponse,
        include_charts: bool,
        include_raw_data: bool
    ) -> bytes:
        """
        Export report to PDF format

        In production, use a PDF generation library like:
        - ReportLab
        - WeasyPrint
        - pdfkit
        """
        # Mock PDF generation
        pdf_content = f"""
        PDF REPORT: {report.name}
        Generated: {report.generated_at}
        Date Range: {report.date_range}

        Sections: {len(report.sections)}

        [This is a mock PDF. In production, use ReportLab or WeasyPrint]
        """

        return pdf_content.encode('utf-8')

    async def _export_to_excel(
        self,
        report: GeneratedReportResponse,
        include_charts: bool,
        include_raw_data: bool
    ) -> bytes:
        """
        Export report to Excel format

        In production, use openpyxl or xlsxwriter:
        - Create workbook
        - Add sheets for each section
        - Add charts if include_charts=True
        - Format cells and add styling
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, Reference
            from openpyxl.styles import Font, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"

            # Add report header
            ws['A1'] = report.name
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M')}"
            ws['A3'] = f"Date Range: {report.date_range}"

            row = 5

            # Add each section
            for section in report.sections:
                ws[f'A{row}'] = section.title
                ws[f'A{row}'].font = Font(size=14, bold=True)
                row += 2

                # Add data headers
                ws[f'A{row}'] = "Date"
                ws[f'B{row}'] = "Value"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                row += 1

                # Add data points
                start_row = row
                for point in section.data:
                    ws[f'A{row}'] = point.date
                    ws[f'B{row}'] = point.value
                    row += 1

                # Add chart if requested
                if include_charts and len(section.data) > 1:
                    chart = LineChart()
                    chart.title = section.title
                    chart.y_axis.title = str(section.metric.value)

                    data = Reference(ws, min_col=2, min_row=start_row-1, max_row=row-1)
                    cats = Reference(ws, min_col=1, min_row=start_row, max_row=row-1)

                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)

                    ws.add_chart(chart, f'D{start_row}')

                # Add summary stats if raw data requested
                if include_raw_data and section.summary_stats:
                    row += 1
                    ws[f'A{row}'] = "Summary Statistics"
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1

                    for key, value in section.summary_stats.items():
                        ws[f'A{row}'] = key.replace('_', ' ').title()
                        ws[f'B{row}'] = value
                        row += 1

                row += 2

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        except ImportError:
            # Fallback if openpyxl not available
            mock_excel = f"Excel Report: {report.name}\n[Install openpyxl for actual Excel generation]"
            return mock_excel.encode('utf-8')

    async def _export_to_csv(
        self,
        report: GeneratedReportResponse
    ) -> bytes:
        """
        Export report to CSV format

        Creates a simple CSV with all data from all sections
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([report.name])
        writer.writerow([f"Generated: {report.generated_at}"])
        writer.writerow([f"Date Range: {report.date_range}"])
        writer.writerow([])

        # Write each section
        for section in report.sections:
            writer.writerow([section.title])
            writer.writerow(["Date", "Value", "Label"])

            for point in section.data:
                writer.writerow([point.date, point.value, point.label or ""])

            # Add summary stats
            if section.summary_stats:
                writer.writerow([])
                writer.writerow(["Summary Statistics"])
                for key, value in section.summary_stats.items():
                    writer.writerow([key.replace('_', ' ').title(), value])

            writer.writerow([])
            writer.writerow([])

        return output.getvalue().encode('utf-8')

    async def get_export_file(
        self,
        export_id: str,
        user_id: str
    ) -> Optional[Dict[str, Any]]:
        """
        Get an export file for download

        Returns file data and metadata
        """
        export = self.exports_cache.get(export_id)

        if not export:
            return None

        # Check expiration (24 hours)
        if datetime.now() > export["created_at"] + timedelta(hours=24):
            del self.exports_cache[export_id]
            return None

        return export
